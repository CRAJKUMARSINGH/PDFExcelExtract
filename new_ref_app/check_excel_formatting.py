"""
Check Excel Formatting Quality
Analyzes the structure of converted Excel files
"""

import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def analyze_excel_file(file_path):
    """Analyze the structure and formatting of an Excel file"""
    try:
        # Load workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"Analyzing: {Path(file_path).name}")
        print("-" * 40)
        
        # Get dimensions
        print(f"Sheet name: {ws.title}")
        print(f"Max row: {ws.max_row}")
        print(f"Max column: {get_column_letter(ws.max_column)} ({ws.max_column})")
        
        # Check for tables/data structures
        data_rows = 0
        empty_rows = 0
        table_headers_found = 0
        
        for row in range(1, min(30, ws.max_row + 1)):  # Check first 30 rows
            empty_cells = 0
            total_cells = 0
            row_data = []
            
            for col in range(1, min(15, ws.max_column + 1)):  # Check first 15 columns
                cell = ws.cell(row=row, column=col)
                total_cells += 1
                if cell.value is None or str(cell.value).strip() == "":
                    empty_cells += 1
                else:
                    row_data.append(str(cell.value)[:30])  # First 30 chars
                    # Check for table headers
                    if "TABLE" in str(cell.value).upper() and "PAGE" in str(cell.value).upper():
                        table_headers_found += 1
            
            if empty_cells == total_cells:
                empty_rows += 1
            else:
                data_rows += 1
                if row <= 15:  # Show first 15 data rows
                    print(f"  Row {row}: {' | '.join(row_data)}")
        
        print(f"\nStructure analysis:")
        print(f"  Data rows sampled: {data_rows}")
        print(f"  Empty rows sampled: {empty_rows}")
        print(f"  Table headers found: {table_headers_found}")
        
        # Check if we have multi-column data
        if ws.max_column > 1:
            print(f"  Multi-column data: YES ({ws.max_column} columns)")
        else:
            print(f"  Multi-column data: NO ({ws.max_column} column)")
        
        # Check column widths
        print(f"\nColumn widths (first 6 columns):")
        for col in range(1, min(7, ws.max_column + 1)):
            col_letter = get_column_letter(col)
            width = ws.column_dimensions[col_letter].width
            print(f"  Column {col_letter}: {width}")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"Error analyzing {file_path}: {e}")
        return False

def main():
    """Main function"""
    print("IMPROVED Excel Formatting Analysis")
    print("=" * 35)
    
    # Look for Excel files in improved_output directory
    output_dir = "improved_output"
    
    if not os.path.exists(output_dir):
        print(f"Directory {output_dir} not found")
        return
    
    # Find Excel files
    excel_files = []
    for file in os.listdir(output_dir):
        if file.endswith('_TABLES.xlsx'):
            excel_files.append(os.path.join(output_dir, file))
    
    if not excel_files:
        print("No Excel files found in improved_output directory")
        return
    
    print(f"Found {len(excel_files)} Excel files")
    print()
    
    # Analyze files with detected tables
    analyzed_files = 0
    for excel_file in excel_files:
        # Check if this file had tables detected
        filename = Path(excel_file).name
        if "PADLI MANOHAR" in filename or "bill bridge" in filename or "RETWALL" in filename or "SLAB" in filename or "r1" in filename or "r2" in filename or "r3" in filename or "r4" in filename:
            analyze_excel_file(excel_file)
            print()
            analyzed_files += 1
            
            if analyzed_files >= 3:  # Limit to first 3 files with tables
                break
    
    if analyzed_files == 0:
        # Analyze first few files if no tables found
        for i, excel_file in enumerate(excel_files[:2]):
            analyze_excel_file(excel_file)
            print()

if __name__ == "__main__":
    main()